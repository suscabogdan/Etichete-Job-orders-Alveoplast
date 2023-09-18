import openpyxl
import os
import datetime
import math
from copy import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as openpyxlImage
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils.cell import coordinate_from_string
from io import BytesIO
from PIL import Image as PILImage
from PIL import Image, ImageDraw

def get_top_left_cell_of_merged_region(worksheet, cell_address):
    """Identify the top-left cell of a merged region."""
    for merged_range in worksheet.merged_cells.ranges:
        if cell_address in merged_range:
            return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return worksheet[cell_address]

# Ensure the "New Job Orders" directory exists
# output_dir = r"C:\Users\gabri\Desktop\Job Orders Tests\New Job Orders"
output_dir = r"L:\1_EXTRUDARE\01. PLANIFICARE PRODUCTIE\1. CALENDAR COMENZI PRODUCTIE\Script&ETICHETE\Job Orders\New Job Orders"
# output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "New Job Orders")
if not os.path.exists(output_dir):
    os.mkdir(output_dir)

# Load the main data workbook
source_wb = openpyxl.load_workbook("EVIDENTA COMANDA ALVEOPLAST.xlsx")
source_ws = source_wb["COMENZI ALVEOPLAST"]

start_row = int(input("Enter the starting row: "))
end_row = int(input("Enter the ending row: "))
# start_row = 847
# end_row = 848

def copy_range(src_ws, dest_ws, src_range, dest_cell):    
    rows = src_ws[src_range]
    dest_cell = dest_ws[dest_cell]
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            target_cell = dest_cell.offset(i, j)
            top_left_target_cell = get_top_left_cell_of_merged_region(dest_ws, target_cell.coordinate)
            top_left_target_cell.value = cell.value
            if cell.has_style:
                top_left_target_cell._style = copy(cell._style)

# Create an array of images
images = ["Images/Alveoplast.png", "Images/Energie Verde.jpeg", "Images/PP.png", "Images/SARC.png"]
anchor_a = ['AG42', 'T47', 'AD20', 'AG18']
anchor_b = ['AX42', 'AK47', 'AU20', 'AX18']

# Mapping for new Job Oreder Template file
# Mapping for Job B
mapping = {
    "H": "K49", # Denumire client
    "I": "K51", # Cod produs
    "B": "K50", # Numar comanda client
    "P": ["K16", "K53"], # Cantitate de extrudat
    # "X": "J17", # Greutate totala de extrudat
    "T": "K4", # Lungime
    "U": "K5", # Latime
    "W": "J9", # Densitate
    "V": "J7", # Grosime
    "S": "J11", # Culoare
    "M": "K19", # Coli / palet
    # "O": "N24",  # Numar paleti (P/M)
    "R": "M23" # Cod reteta
}

# Mapping for Job A
mapping_job_a = {
    "H": "O49", # Denumire client
    "I": "O51", # Cod produs
    "B": "O50", # Numar comanda client
    "P": ["O16", "O53"], # Cantitate de extrudat
    # "X": "J17", # Greutate totala de extrudat
    "T": "O4", # Lungime
    "U": "O5", # Latime
    "W": "J9", # Densitate
    "V": "J7", # Grosime
    "S": "J11", # Culoare
    "M": "O19", # Coli / palet
    # "O": "N24",  # Numar paleti (P/M)
    "R": "M23" # Cod reteta
}

def load_r_mapping_from_file(filename, color_value):
    r_mapping = {}
    with open(filename, 'r') as file:
        current_key = None
        for line in file:
            line = line.strip()
            if not line:
                continue
            if line.endswith(':'):
                current_key = line[:-1]
                r_mapping[current_key] = {}
            else:
                cell, value = line.split('=', 1)
                # Replace the color placeholder with the actual color value
                value = value.replace('{color}', color_value)
                r_mapping[current_key][cell.strip()] = value.strip()
    return r_mapping

def sanitize_filename(filename):
    """Sanitize filenames."""
    sanitized_name = ''.join(char if char.isalnum() or char in [' ', '_'] else '-' for char in filename)
    return sanitized_name

def populate_and_save_template(job_a_row, job_b_row, num_pallets_a, remaining_sheets_a, num_pallets_b=None, remaining_sheets_b=None):
    if num_pallets_b is None:
        num_pallets_b = 0
    if remaining_sheets_b is None:
        remaining_sheets_b = 0
    
    template_wb = openpyxl.load_workbook('JO&EP Template 2.xlsx')
    template_ws = template_wb["SABLON"]

    for col, dest_cells in mapping_job_a.items():
        if not isinstance(dest_cells, list):
            dest_cells = [dest_cells]
        for dest_cell in dest_cells:
            top_left_cell = get_top_left_cell_of_merged_region(template_ws, dest_cell)
        
            # Default logic
            source_value = source_ws[col + str(job_a_row)].value
            top_left_cell.value = source_value

    if job_b_row:  # Only proceed if there's a B job
        for col, dest_cells in mapping.items():
            if not isinstance(dest_cells, list):
                dest_cells = [dest_cells]
            for dest_cell in dest_cells:
                top_left_cell = get_top_left_cell_of_merged_region(template_ws, dest_cell)
            
                # Default logic
                source_value = source_ws[col + str(job_b_row)].value
                top_left_cell.value = source_value

    # Mapping for values in column R
    color_value = source_ws["S" + str(job_a_row)].value

    r_mapping = load_r_mapping_from_file("Retete.txt", color_value)

    # Check value in column R for job A
    r_value_a = source_ws["R" + str(job_a_row)].value
    if r_value_a in r_mapping:
        for cell, value in r_mapping[r_value_a].items():
            top_left_cell = get_top_left_cell_of_merged_region(template_ws, cell)
            top_left_cell.value = value


    # If there's a B job, also check its R value
    if job_b_row:
        r_value_b = source_ws["R" + str(job_b_row)].value
        if r_value_b in r_mapping:
            for cell, value in r_mapping[r_value_b].items():
                top_left_cell = get_top_left_cell_of_merged_region(template_ws, cell)
                top_left_cell.value = value


    total_sheets_a = int(source_ws.cell(row=job_a_row, column=16).value)
    total_sheets_b = int(source_ws.cell(row=job_b_row, column=16).value) if job_b_row else 0

    # Copy and paste the Job B template the correct number of times
    if job_b_row:
        for i in range(num_pallets_b):
            # Copy the Job A template
            src_range = "AI1:AY56"
            dest_cell = "AI" + str(59 + i * 57)
            copy_range(template_ws, template_ws, src_range, dest_cell)

    # Write the correct information into the Job B labels
    if job_b_row:
        for i in range(num_pallets_b):
            row_v79 = 104 + i * 57
            row_AA79 = 104 + i * 57

            template_ws.cell(row=row_v79, column=42).value = i + 1
            template_ws.cell(row=row_AA79, column=46).value = num_pallets_b

    # Copy and paste the Job A template the correct number of times
    if job_a_row and current_a_row:
        for i in range(num_pallets_a):
            # Copy the Job B template
            src_range = "R1:AH56"
            dest_cell = "R" + str(59 + i * 57)
            copy_range(template_ws, template_ws, src_range, dest_cell)

    # Write the correct information into the Job A labels
    if job_a_row and current_a_row:
        for i in range(num_pallets_a):
            row_ah79 = 104 + i * 57
            row_AL79 = 104 + i * 57

            template_ws.cell(row=row_ah79, column=25).value = i + 1
            template_ws.cell(row=row_AL79, column=29).value = num_pallets_a
    
    cell_interval = 57

    for i in range(num_pallets_a + 1):
        for j, (img_name, anchor) in enumerate(zip(images, anchor_a)):
            # Compute the anchor for this repetition
            col, row = coordinate_from_string(anchor)
            row += cell_interval * i
            new_anchor = f'{col}{row}'
        
            # Add image
            img = openpyxlImage(img_name)
        
            if j == 0:
                img.width = int(63 * 1.33)
                img.height = int(163 * 1.33)
            elif j == 1:
                img.width = int(134 * 1.33)
                img.height = int(77 * 1.33)
            elif j == 2:
                img.width = int(94 * 1.33)
                img.height = int(107 * 1.33)
            elif j == 3:
                img.width = int(58 * 1.33)
                img.height = int(474 * 1.33)
        
            img.anchor = new_anchor
            template_ws.add_image(img)

        if job_a_row and current_b_row:
            for i in range(num_pallets_b + 1):
                for j, (img_name, anchor) in enumerate(zip(images, anchor_b)):
                    # Compute the anchor for this repetition
                    col, row = coordinate_from_string(anchor)
                    row += cell_interval * i
                    new_anchor = f'{col}{row}'
        
                    # Add image
                    img = openpyxlImage(img_name)
        
                    if j == 0:
                        img.width = int(63 * 1.33)
                        img.height = int(163 * 1.33)
                    elif j == 1:
                        img.width = int(134 * 1.33)
                        img.height = int(77 * 1.33)
                    elif j == 2:
                        img.width = int(94 * 1.33)
                        img.height = int(107 * 1.33)
                    elif j == 3:
                        img.width = int(58 * 1.33)
                        img.height = int(474 * 1.33)
        
                    img.anchor = new_anchor
                    template_ws.add_image(img)



    # Code for naming and saving the file
    client = source_ws["H" + str(job_a_row)].value or "Unknown"
    length = source_ws["S" + str(job_a_row)].value or "Unknown"
    width = source_ws["T" + str(job_a_row)].value or "Unknown"
    density = source_ws["V" + str(job_a_row)].value or "Unknown"
    thickness = source_ws["U" + str(job_a_row)].value or "Unknown"
    filename = f"JO&EP - {client}-{length}-{width}-{thickness}-{density}"
    if job_b_row:
        client = source_ws["H" + str(job_b_row)].value or "Unknown"
        length = source_ws["S" + str(job_b_row)].value or "Unknown"
        width = source_ws["T" + str(job_b_row)].value or "Unknown"
        filename += f"-{client}-{length}-{width}"
    filename = sanitize_filename(filename) + ".xlsx"
    filepath = os.path.join(os.path.dirname(__file__), output_dir, filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    counter = 1
    base_filename = filepath
    while os.path.exists(filepath):
        filename = base_filename.replace(".xlsx", f" ({counter}).xlsx")
        filepath = os.path.join(os.path.dirname(__file__), output_dir, filename)
        counter += 1
    template_wb.save(filepath)
    # print(f"Saved: {filename}")
    print(f"Saved {filepath}")

# Main Loop Logic
current_a_row = None
current_b_row = None
consecutive_a_rows = []

for row in range(start_row, end_row + 1):
    job_type = source_ws["Q" + str(row)].value
    current_b_row = row
    if job_type == 'A':
        current_a_row = row
        # Calculate the number of pallets and remaining sheets for Job A
        source_value_p = source_ws["P" + str(current_a_row)].value
        source_value_m = source_ws["M" + str(current_a_row)].value
        num_pallets_a = math.ceil(source_value_p / source_value_m) if source_value_m else 0
        remaining_sheets_a = source_value_p - num_pallets_a * source_value_m
        # Save a separate file for the 'A' job
        populate_and_save_template(current_a_row, None, num_pallets_a, remaining_sheets_a)
    elif job_type == 'B' and current_a_row:
        # Calculate the number of pallets and remaining sheets for Job B
        source_value_p = source_ws["P" + str(row)].value
        source_value_m = source_ws["M" + str(row)].value
        num_pallets_b = math.ceil(source_value_p / source_value_m) if source_value_m else 0
        remaining_sheets_b = source_value_p - num_pallets_b * source_value_m
        # Save a separate file for the 'A' job paired with the 'B' job
        populate_and_save_template(current_a_row, current_b_row, num_pallets_a, remaining_sheets_a, num_pallets_b, remaining_sheets_b)

print("All files created successfully!")
